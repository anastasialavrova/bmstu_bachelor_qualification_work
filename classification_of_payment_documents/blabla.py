import openpyxl
import json


if __name__ == '__main__':
    wb = openpyxl.load_workbook('data/table.xlsx')

    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column

    code = []
    name = []
    cnt = 0

    dict = {}

    for i in range(2, rows + 1):
        cnt = cnt + 1;
        for j in range(1, cols + 1):
            if (j == 5):
                cell = sheet.cell(row = i, column = j)
                code.append(str(cell.value))
            if (j == 6):
                cell = sheet.cell(row=i, column=j)
                name.append(str(cell.value))

    for i in range(len(code)):
        if (i != 0):
            if (code[i - 1] != code[i]):
                dict[code[i]] = name[i]
        else:
            dict[code[i]] = name[i]

    with open('data/name_of_code.txt', 'w') as outfile:
        json.dump(dict, outfile, ensure_ascii=False)

    print(dict)




