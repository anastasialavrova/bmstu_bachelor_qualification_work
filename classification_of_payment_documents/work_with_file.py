import openpyxl


def read_data():
    wb = openpyxl.load_workbook('data/table.xlsx')

    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column

    name = []
    description = []
    cnt = 0

    for i in range(2, rows + 1):
        cnt = cnt + 1;
        for j in range(1, cols + 1):
            if (j == 5):
                cell = sheet.cell(row = i, column = j)
                name.append(str(cell.value))
            if (j == 7):
                cell = sheet.cell(row=i, column=j)
                description.append(str(cell.value))

    return name, description

def get_names():
    f = open("names/female_names_rus.txt", "r")
    f2 = open("names/male_names_rus.txt", "r")
    f3 = open("names/male_surnames_rus.txt", "r")
    f4 = open("names/names3.txt", "r")
    female_names = []
    male_names = []
    male_surnames = []
    names = []
    for line in f:
        female_names.append(line[:-1].lower())
    for line in f2:
        male_names.append(line[:-1].lower())
    for line in f3:
        male_surnames.append(line[:-1].lower())
    for line in f4:
        names.append(line[:-1].lower())
    f.close()
    res = female_names + male_names + male_surnames + names
    return res
